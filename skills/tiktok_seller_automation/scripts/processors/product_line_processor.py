import os
import json
import openpyxl
from copy import copy
from datetime import datetime
from skills.tiktok_seller_automation.scripts.processors.base_processor import BaseProcessor

class ProductLineProcessor(BaseProcessor):
    """
    Agent for filling Product_Line_Subscription_Report.xlsx.
    Maintains formatting, adds dated sheets, and calculates WoW.
    """
    def __init__(self, template_path, report_path):
        super().__init__(template_path, report_path)
        self.report_path = report_path # This is the master report file

    def _copy_style(self, source_cell, target_cell):
        """Deep copy style from one cell to another."""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    def _get_previous_sheet_data(self, workbook, current_sheet_name):
        """Extracts ID -> GMV mapping from the chronologically previous sheet."""
        # Sort sheets to find the one before the current_sheet_name
        # Note: Sheet names are YYYY-MM-DD
        sorted_sheets = sorted([s for s in workbook.sheetnames if s != "template"])
        
        try:
            current_idx = sorted_sheets.index(current_sheet_name)
            if current_idx == 0:
                return {} # No previous sheet
            
            prev_sheet_name = sorted_sheets[current_idx - 1]
            prev_sheet = workbook[prev_sheet_name]
            print(f"Detected previous sheet for WoW: {prev_sheet_name}")
            
            data = {}
            for row in range(2, prev_sheet.max_row + 1):
                raw_id = prev_sheet.cell(row=row, column=2).value
                if not raw_id: continue
                
                try:
                    if isinstance(raw_id, float):
                        p_id = format(raw_id, '.0f')
                    else:
                        p_id = str(raw_id).strip()
                    
                    # WoW baseline should be Total GMV (Column 5)
                    gmv = prev_sheet.cell(row=row, column=5).value
                    if gmv is not None and not isinstance(gmv, str): # Skip formulas/strings
                        data[p_id] = float(gmv)
                except:
                    pass
            return data
        except (ValueError, IndexError):
            return {}

    def process(self, sheet_name="template"):
        if not self.raw_data:
            raise ValueError("No raw data loaded.")

        # 1. Load or Create the Master Report
        if os.path.exists(self.report_path):
            wb_report = openpyxl.load_workbook(self.report_path)
        else:
            wb_report = openpyxl.Workbook()
            if "Sheet" in wb_report.sheetnames:
                del wb_report["Sheet"]

        # 2. Load Template
        wb_temp = openpyxl.load_workbook(self.template_path)
        temp_sheet = wb_temp[sheet_name] if sheet_name in wb_temp.sheetnames else wb_temp.active

        # 3. Create New Sheet named by Date
        report_date = self.raw_data.get('metadata', {}).get('start_date', datetime.now().strftime("%Y-%m-%d"))
        if report_date in wb_report.sheetnames:
            print(f"Warning: Sheet for {report_date} already exists. Overwriting...")
            del wb_report[report_date]
        
        new_sheet = wb_report.create_sheet(title=report_date)

        # 4. Copy Template Structure and Styles
        for row in temp_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                self._copy_style(cell, new_cell)
        
        for col_letter, col_dim in temp_sheet.column_dimensions.items():
            new_sheet.column_dimensions[col_letter].width = col_dim.width

        # 5. Prepare Raw Data Lookups
        # Lookup 1: Subscription GMV
        sub_data = self.raw_data.get('sections', {}).get('subscription_data', [])
        sub_gmv_lookup = {}
        for item in sub_data:
            p_id = str(item.get('product_id')).strip()
            gmv_val = item.get('subscription_gmv', {}).get('amount')
            if p_id and gmv_val is not None:
                sub_gmv_lookup[p_id] = float(gmv_val)
        
        # Lookup 2: Total GMV (from General Performance)
        gen_data = self.raw_data.get('sections', {}).get('product_performance_general', [])
        total_gmv_lookup = {}
        for item in gen_data:
            p_id = str(item.get('meta', {}).get('product_id')).strip()
            gmv_val = item.get('stats', {}).get('gmv', {}).get('amount')
            if p_id and gmv_val is not None:
                total_gmv_lookup[p_id] = float(gmv_val)
        
        prev_gmv_lookup = self._get_previous_sheet_data(wb_report, report_date)

        # 6. Fill Data and Formulas
        print(f"Filling data for sheet: {report_date}...")
        match_count = 0
        
        # Pass 1: Products
        total_rows = []
        for row in range(2, new_sheet.max_row + 1):
            product_name = str(new_sheet.cell(row=row, column=1).value or "")
            id_val = new_sheet.cell(row=row, column=2).value
            
            if "Total" in product_name and not id_val:
                total_rows.append(row)
                continue

            if id_val:
                try:
                    if isinstance(id_val, float):
                        p_id = format(id_val, '.0f')
                    else:
                        p_id = str(id_val).strip()
                    
                    # Fill Subscription GMV (Column 3)
                    if p_id in sub_gmv_lookup:
                        new_sheet.cell(row=row, column=3).value = sub_gmv_lookup[p_id]
                    
                    # Fill Total GMV (Column 5)
                    if p_id in total_gmv_lookup:
                        curr_total_gmv = total_gmv_lookup[p_id]
                        new_sheet.cell(row=row, column=5).value = curr_total_gmv
                        match_count += 1
                        
                        # WoW calculation based on Total GMV (now Column 7 after shift)
                        if p_id in prev_gmv_lookup:
                            prev_gmv = prev_gmv_lookup[p_id]
                            if prev_gmv > 0:
                                wow = (curr_total_gmv - prev_gmv) / prev_gmv
                                new_sheet.cell(row=row, column=7).value = wow
                except:
                    pass

        # Pass 2: Totals and Percentages
        # First, find all rows that contain individual products (have an ID)
        product_rows = []
        for row in range(2, new_sheet.max_row + 1):
            if new_sheet.cell(row=row, column=2).value:
                product_rows.append(row)
        
        # We'll use an Excel formula to sum all these specific rows for the Grand Total
        # format: SUM(E[r1], E[r2], ...)
        grand_total_formula = "=SUM(" + ",".join([f"E{r}" for r in product_rows]) + ")"
        
        # We can place this formula in a temporary cell or just use it in the percentage formula
        # To keep it clean, let's just use the logic in each percentage cell
        
        group_start = 2
        for total_row in total_rows:
            # Column 3: Subscription GMV Total
            sub_sum_range = f"C{group_start}:C{total_row-1}"
            new_sheet.cell(row=total_row, column=3).value = f"=SUM({sub_sum_range})"
            
            # Column 5: Total GMV Total
            tot_sum_range = f"E{group_start}:E{total_row-1}"
            new_sheet.cell(row=total_row, column=5).value = f"=SUM({tot_sum_range})"
            
            # Percentage Formulas for all products in this group
            for row in range(group_start, total_row):
                # Column 4: % of Sub GMV = C[row] / E[row] (Sub GMV / Total GMV for this product)
                new_sheet.cell(row=row, column=4).value = f"=IF(E{row}>0, C{row}/E{row}, 0)"
                
                # Column 6: % of Total GMV = E[row] / SUM(all product rows)
                # Note: We use absolute references for the product rows in the SUM
                new_sheet.cell(row=row, column=6).value = f"=IF({grand_total_formula.replace('=', '')}>0, E{row}/({grand_total_formula.replace('=', '')}), 0)"
            
            group_start = total_row + 1

        print(f"Successfully matched and filled {match_count} products.")

        # 7. Save the Master Report
        wb_report.save(self.report_path)
        print(f"Master report updated: {self.report_path}")
        return self.report_path
