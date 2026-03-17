import os
import json
import openpyxl
from copy import copy
from datetime import datetime
from .base_processor import BaseProcessor

class ProductLineProcessor(BaseProcessor):
    def __init__(self, template_path, report_path):
        super().__init__(template_path, report_path)
        self.report_path = report_path

    def _copy_style(self, source_cell, target_cell):
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

    def _get_previous_sheet_data(self, workbook, current_sheet_name):
        """Extracts metrics from the previous sheet, calculating values for formula cells."""
        # Filter for valid dated sheets (YYYY-MM-DD)
        valid_sheets = []
        for s in workbook.sheetnames:
            if s == "template": continue
            try:
                datetime.strptime(s, "%Y-%m-%d")
                valid_sheets.append(s)
            except ValueError:
                continue
        
        sorted_sheets = sorted(valid_sheets)
        
        try:
            # Find the latest sheet that is strictly before the current_sheet_name
            past_sheets = [s for s in sorted_sheets if s < current_sheet_name]
            if not past_sheets:
                print(f"No previous dated sheets found before {current_sheet_name}.")
                return {"products": {}, "aggregates": {}}
            
            prev_sheet_name = past_sheets[-1]
            prev_sheet = workbook[prev_sheet_name]
            print(f"Detected previous sheet for WoW: {prev_sheet_name}")
            
            product_data = {}
            aggregate_data = {}
            current_group_metrics = []
            
            for row in range(2, prev_sheet.max_row + 1):
                product_name = str(prev_sheet.cell(row=row, column=1).value or "").strip()
                raw_id = prev_sheet.cell(row=row, column=2).value
                
                # Full list of metric columns to capture for WoW:
                # GMV(5), Items(8), VidGMV(10), VidGMV%(12), Orders(14), VidImp(17), VidCTR(19), VidCVR(21), PCardGMV(23), PCardImp(25), PCardCTR(27), PCardCVR(29)
                metric_cols = [3, 5, 8, 10, 12, 14, 17, 19, 21, 23, 25, 27, 29]
                metrics = {}
                for col in metric_cols:
                    val = prev_sheet.cell(row=row, column=col).value
                    try:
                        metrics[col] = float(val) if val is not None and not isinstance(val, str) else 0.0
                    except:
                        metrics[col] = 0.0

                if raw_id:
                    p_id = format(raw_id, '.0f') if isinstance(raw_id, float) else str(raw_id).strip()
                    # Re-calculate derived column 12 if it's 0/missing (likely due to it being a formula)
                    if metrics.get(5, 0) > 0:
                        metrics[12] = metrics.get(10, 0) / metrics[5]
                    
                    product_data[p_id] = metrics
                    current_group_metrics.append(metrics)
                elif "Total" in product_name:
                    # Sum basics for group
                    agg_metrics = {col: sum(m[col] for m in current_group_metrics) for col in [3, 5, 8, 10, 14, 17, 23, 25]}
                    # Average CTR/CVR for group
                    for col in [19, 21, 27, 29]:
                        agg_metrics[col] = sum(m[col] for m in current_group_metrics) / len(current_group_metrics) if current_group_metrics else 0.0
                    # Derived metrics
                    agg_metrics[12] = agg_metrics[10] / agg_metrics[5] if agg_metrics[5] > 0 else 0.0
                    
                    aggregate_data[product_name] = agg_metrics
                    current_group_metrics = []
                    
            return {"products": product_data, "aggregates": aggregate_data}
        except (ValueError, IndexError):
            return {"products": {}, "aggregates": {}}

    def process(self, sheet_name="template"):
        if not self.raw_data:
            raise ValueError("No raw data loaded.")

        report_date = self.raw_data.get('metadata', {}).get('start_date', datetime.now().strftime("%Y-%m-%d"))
        if os.path.exists(self.report_path):
            wb_report = openpyxl.load_workbook(self.report_path)
            prev_data_lookup = self._get_previous_sheet_data(wb_report, report_date)
        else:
            wb_report = openpyxl.Workbook()
            if "Sheet" in wb_report.sheetnames: del wb_report["Sheet"]
            prev_data_lookup = {"products": {}, "aggregates": {}}

        wb_temp = openpyxl.load_workbook(self.template_path)
        temp_sheet = wb_temp[sheet_name] if sheet_name in wb_temp.sheetnames else wb_temp.active

        if report_date in wb_report.sheetnames:
            print(f"Warning: Overwriting existing sheet {report_date}")
            del wb_report[report_date]
        new_sheet = wb_report.create_sheet(title=report_date)

        for row in temp_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                self._copy_style(cell, new_cell)
        for col_letter, col_dim in temp_sheet.column_dimensions.items():
            new_sheet.column_dimensions[col_letter].width = col_dim.width

        # Lookups
        sub_data = self.raw_data.get('sections', {}).get('subscription_data', [])
        sub_gmv_lookup = {str(i.get('product_id')).strip(): float(i.get('subscription_gmv', {}).get('amount', 0)) for i in sub_data}
        
        gen_data = self.raw_data.get('sections', {}).get('product_performance_general', [])
        stats_lookup = {}
        for item in gen_data:
            p_id = str(item.get('meta', {}).get('product_id')).strip()
            if p_id:
                s = item.get('stats', {})
                stats_lookup[p_id] = {
                    5: float(s.get('gmv', {}).get('amount', 0)),
                    8: float(s.get('unit_sold_cnt', 0)),
                    10: float(s.get('video_gmv', {}).get('amount', 0)),
                    14: float(s.get('order_cnt', 0)),
                    17: float(s.get('video_listing_impression_cnt', 0)),
                    19: float(s.get('video_ctr', 0)),
                    21: float(s.get('video_cvr', 0)),
                    23: float(s.get('product_card_gmv', {}).get('amount', 0)),
                    25: float(s.get('product_card_listing_impression_cnt', 0)),
                    27: float(s.get('product_card_ctr', 0)),
                    29: float(s.get('product_card_cvr', 0))
                }

        # Fill Data
        print(f"Filling data for sheet: {report_date}...")
        total_rows_meta = [] 
        product_rows = []
        
        for row in range(2, new_sheet.max_row + 1):
            product_name = str(new_sheet.cell(row=row, column=1).value or "").strip()
            id_val = new_sheet.cell(row=row, column=2).value
            
            if "Total" in product_name and not id_val:
                total_rows_meta.append((row, product_name))
                continue

            if id_val:
                try:
                    p_id = format(id_val, '.0f') if isinstance(id_val, float) else str(id_val).strip()
                    product_rows.append(row)
                    if p_id in sub_gmv_lookup:
                        new_sheet.cell(row=row, column=3).value = sub_gmv_lookup[p_id]
                    
                    if p_id in stats_lookup:
                        cur = stats_lookup[p_id]
                        prev = prev_data_lookup["products"].get(p_id, {})
                        
                        # [Metric Col, WoW Col]
                        mappings = [[5, 7], [8, 9], [10, 11], [14, 15], [17, 18], [19, 20], [21, 22], [23, 24], [25, 26], [27, 28], [29, 30]]
                        for m_col, wow_col in mappings:
                            new_sheet.cell(row=row, column=m_col).value = cur[m_col]
                            if m_col in prev and prev[m_col] > 0:
                                new_sheet.cell(row=row, column=wow_col).value = (cur[m_col] - prev[m_col]) / prev[m_col]

                        # Formulas
                        new_sheet.cell(row=row, column=16).value = f"=IF(N{row}>0, E{row}/N{row}, 0)" 
                        new_sheet.cell(row=row, column=12).value = f"=IF(E{row}>0, J{row}/E{row}, 0)" 
                        
                        # Video % WoW (13)
                        cur_gmv = cur.get(5, 0)
                        cur_vid_gmv = cur.get(10, 0)
                        this_pct = cur_vid_gmv / cur_gmv if cur_gmv > 0 else 0
                        
                        prev_vid_pct = prev.get(12)
                        if prev_vid_pct is not None and prev_vid_pct > 0:
                            new_sheet.cell(row=row, column=13).value = (this_pct - prev_vid_pct) / prev_vid_pct
                except Exception as e:
                    print(f"Error processing row {row}: {e}")

        # Totals and Percentages
        grand_total_formula = "SUM(" + ",".join([f"E{r}" for r in product_rows]) + ")"
        group_start = 2
        for total_row, total_name in total_rows_meta:
            prev_agg = prev_data_lookup["aggregates"].get(total_name, {})
            
            # Sum columns: 3, 5, 8, 10, 14, 17, 23, 25
            agg_sum_cols = [3, 5, 8, 10, 14, 17, 23, 25]
            # WoW Map
            agg_wow_map = {5: 7, 8: 9, 10: 11, 14: 15, 17: 18, 19: 20, 21: 22, 23: 24, 25: 26, 27: 28, 29: 30}
            
            cur_agg_vals = {col: 0.0 for col in [3, 5, 8, 10, 14, 17, 19, 21, 23, 25, 27, 29]}
            p_count = 0
            for r in range(group_start, total_row):
                rid_val = new_sheet.cell(row=r, column=2).value
                if rid_val:
                    r_pid = format(rid_val, '.0f') if isinstance(rid_val, float) else str(rid_val).strip()
                    p_count += 1
                    for col in cur_agg_vals:
                        if col == 3: cur_agg_vals[col] += sub_gmv_lookup.get(r_pid, 0)
                        else: cur_agg_vals[col] += stats_lookup.get(r_pid, {}).get(col, 0)
            
            if p_count > 0:
                for c in [19, 21, 27, 29]: cur_agg_vals[c] /= p_count

            for col in agg_sum_cols:
                col_letter = openpyxl.utils.get_column_letter(col)
                new_sheet.cell(row=total_row, column=col).value = f"=SUM({col_letter}{group_start}:{col_letter}{total_row-1})"
            
            for m_col, wow_col in agg_wow_map.items():
                if m_col in prev_agg and prev_agg[m_col] > 0:
                    new_sheet.cell(row=total_row, column=wow_col).value = (cur_agg_vals[m_col] - prev_agg[m_col]) / prev_agg[m_col]

            # Aggregate Derived
            new_sheet.cell(row=total_row, column=12).value = f"=IF(E{total_row}>0, J{total_row}/E{total_row}, 0)" # Video %
            this_agg_pct = cur_agg_vals[10] / cur_agg_vals[5] if cur_agg_vals[5] > 0 else 0
            if 12 in prev_agg and prev_agg[12] > 0:
                new_sheet.cell(row=total_row, column=13).value = (this_agg_pct - prev_agg[12]) / prev_agg[12] # Video % WoW

            # Averages for Aggregates
            for c_id, c_letter in {19: 'S', 21: 'U', 27: 'AA', 29: 'AC'}.items():
                new_sheet.cell(row=total_row, column=c_id).value = f"=AVERAGE({c_letter}{group_start}:{c_letter}{total_row-1})"

            # Relative Percentages
            for row in range(group_start, total_row + 1):
                new_sheet.cell(row=row, column=4).value = f"=IF(E{row}>0, C{row}/E{row}, 0)" 
                new_sheet.cell(row=row, column=6).value = f"=IF({grand_total_formula}>0, E{row}/({grand_total_formula}), 0)"
            
            group_start = total_row + 1

        wb_report.save(self.report_path)
        print(f"Master report fully updated with ALL metrics: {self.report_path}")
        return self.report_path
